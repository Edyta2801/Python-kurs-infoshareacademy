import openpyxl

plik = openpyxl.load_workbook("example.xlsx")
# print(type(plik))

# lista arkuszy
arkusze = plik.get_sheet_names()
print(arkusze)

# konkretny arkusz
arkusz = plik.get_sheet_by_name('Owoce')
print("Nazwa arkusza:", arkusz.title)

#aktywny arkusz
print("Aktywny arkusz:", plik.active)

# komórki
komorka = arkusz['A1']
print(komorka)
print(komorka.value)

# koordynaty komorki
print("Adres komorki:", komorka.coordinate)
print("Kolumna komórki:", komorka.column)
print("Wiersz komórki:", komorka.row)

# konkretna komorka z arkusza
print(arkusz.cell(column=2, row=1))
print(arkusz.cell(column=3, row=4))

# rozmiar arkusza
print("Ostatnia kolumna:", arkusz.max_column)
print("Ostatni wiersz:", arkusz.max_row)

# zamiana liter na cyfry w kolumnach
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(56))
print(get_column_letter(905))
print(column_index_from_string('AA'))
print(column_index_from_string('AAG'))

# zakresy danych
print(arkusz['A1':'C3'])

for wiersz in arkusz['A1':'C3']:
    for kom in wiersz:
        print('Komórks {} ma wartość {}'.format(kom.coordinate, kom.value))
    print("-------koniec wiersza---------")



plik.close()